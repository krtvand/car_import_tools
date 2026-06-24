"""Export the SQLite listings table to an .xlsx file (opens in Excel/Numbers)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import db
from models import CarListing

# Column order for the spreadsheet (subset/ordering of the model fields).
COLUMNS = [
    "ad_id", "title", "price", "currency", "year", "mileage_km", "fuel_type",
    "gearbox", "body_type", "engine_size", "power_hp", "colour", "doors",
    "seats", "drive", "mot_till", "availability", "location", "posted_raw",
    "photo_count", "extras", "url", "image_url", "first_seen_at", "last_seen_at",
]


def export_xlsx(path: str | Path = "bazaraki_cars.xlsx") -> Path:
    path = Path(path)
    listings: list[CarListing] = db.all_listings()

    wb = Workbook()
    ws = wb.active
    ws.title = "Cars"

    header_font = Font(bold=True)
    for col_idx, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
    ws.freeze_panes = "A2"

    for row_idx, listing in enumerate(listings, start=2):
        for col_idx, name in enumerate(COLUMNS, start=1):
            value = getattr(listing, name)
            if hasattr(value, "tzinfo"):  # datetime -> naive for Excel
                value = value.replace(tzinfo=None)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Reasonable column widths.
    for col_idx, name in enumerate(COLUMNS, start=1):
        width = 12 if name not in {"title", "url", "image_url", "extras", "location"} else 40
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(path)
    return path.resolve()