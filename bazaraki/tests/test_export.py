"""Tests for the xlsx export."""
from __future__ import annotations

import pytest
from openpyxl import load_workbook
from sqlmodel import create_engine

from bazaraki import db
from bazaraki import export


@pytest.fixture
def temp_db(tmp_path, monkeypatch):
    engine = create_engine(f"sqlite:///{tmp_path / 'test.db'}")
    monkeypatch.setattr(db, "_engine", engine)
    db.init_db()
    return engine


def test_export_creates_file_with_header_and_rows(temp_db, tmp_path):
    db.upsert_listing({"ad_id": 1, "title": "Car A", "url": "u1", "price": 100.0, "year": 2020})
    db.upsert_listing({"ad_id": 2, "title": "Car B", "url": "u2", "price": 200.0, "year": 2021})

    out = export.export_xlsx(tmp_path / "out.xlsx")
    assert out.exists()

    ws = load_workbook(out).active
    assert ws.title == "Cars"
    assert ws.max_row == 3  # header + 2 rows

    header = [c.value for c in ws[1]]
    assert header == export.COLUMNS
    assert header[0] == "ad_id"

    # Values land in the right columns and keep their types.
    assert ws.cell(row=2, column=header.index("ad_id") + 1).value == 1
    assert ws.cell(row=2, column=header.index("price") + 1).value == 100.0
    assert ws.cell(row=2, column=header.index("year") + 1).value == 2020
    assert ws.cell(row=3, column=header.index("title") + 1).value == "Car B"


def test_export_empty_db_writes_header_only(temp_db, tmp_path):
    out = export.export_xlsx(tmp_path / "empty.xlsx")
    ws = load_workbook(out).active
    assert ws.max_row == 1
    assert [c.value for c in ws[1]] == export.COLUMNS


def test_export_datetimes_are_timezone_naive(temp_db, tmp_path):
    # Excel can't store tz-aware datetimes; export must strip tzinfo.
    db.upsert_listing({"ad_id": 1, "title": "A", "url": "u"})
    out = export.export_xlsx(tmp_path / "dt.xlsx")
    ws = load_workbook(out).active
    col = export.COLUMNS.index("first_seen_at") + 1
    value = ws.cell(row=2, column=col).value
    assert value is not None
    assert getattr(value, "tzinfo", None) is None